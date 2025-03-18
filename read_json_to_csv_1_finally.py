#The images processed again with Labelme can be used to identify the animal's weight through an algorithm, and the results can be output in CSV format

import json
import math
import os
import numpy as np
import openpyxl


def get_head_body_length(xmax, xmin, ymax, ymin):
    head_body_length_in_pixel = math.sqrt((xmax - xmin) ** 2 + (ymax - ymin) ** 2)
    # print(head_body_length_in_pixel)
    flight_height = 100
    ground_sampling_distance = 0.00027412 * flight_height
    head_body_length_in_pixel = ground_sampling_distance * head_body_length_in_pixel
    # print(head_body_length_in_pixel)
    return head_body_length_in_pixel


def get_animal_weight(animal_class, head_body_length):
    if animal_class == "cattle":
        # animal_weight=17.308*math.exp(1.3574*head_body_length)
        # #y  =  17.308e1.3574x;  y  =  42.441  x  ^2.6904;  y  =  376.5x  -  454.09
        animal_weight = 376.5 * head_body_length - 454.09  # y  =  17.308e1.3574x;  y  =  42.441  x  ^2.6904;  y  =  376.5x  -  454.09
        if animal_weight < 100:  # 设置牛最小体重为100kg
            animal_weight = 100
    if animal_class == "sheep":
        animal_weight = 55.523 * head_body_length - 34.042  # y  =  55.523x  -  34.042
        if animal_weight < 15:  # 设置羊最小体重为100kg
            animal_weight = 15

    return animal_weight


sheep_flag = 0
cattle_flag = 0
# filepath = "DJI_0311.json"

path = []


def file_name_find(file_dir):
    for root, dirs, files in os.walk(file_dir):
        for j in files:
                if j.endswith('.json'):
                    path.append(str(j))


#json_path = "/Users/yangjianwei/Desktop/json"
json_path ="D:/sheep"
print(json_path)
file_name_find(json_path)
print(path)
for filepath in path:
    with open(json_path + "/" + filepath, 'r') as f:
        work_book = openpyxl.Workbook()
        work_sheet = work_book.create_sheet("Sheet1")
        work_sheet.cell(row=1, column=1).value = "image_name"
        work_sheet.cell(row=1, column=2).value = "class_name"
        work_sheet.cell(row=1, column=3).value = "ymin"
        work_sheet.cell(row=1, column=4).value = "ymax"
        work_sheet.cell(row=1, column=5).value = "xmin"
        work_sheet.cell(row=1, column=6).value = "xmax"
        work_sheet.cell(row=1, column=7).value = "head_body_length"
        work_sheet.cell(row=1, column=8).value = "animal_weight"
        work_sheet.cell(row=1, column=9).value = "sheep_number"
        work_sheet.cell(row=1, column=10).value = "cattle_number"

        tmp = json.loads(f.read())
        print(len(tmp["shapes"]))
        for i in range(len(tmp["shapes"])):
            animal_class = tmp["shapes"][i]["label"]
            if animal_class == "sheep":
                sheep_flag += 1
            else:
                cattle_flag += 1

            arr = tmp["shapes"][i]["points"]
            xmin, ymin = np.amin(arr, axis=0)
            xmax, ymax = np.amax(arr, axis=0)
            head_body_length = get_head_body_length(xmax, xmin, ymax, ymin)
            animal_weight = get_animal_weight(animal_class, head_body_length)

            work_sheet.cell(row=i + 2, column=1).value = filepath[:-5]
            work_sheet.cell(row=i + 2, column=2).value = animal_class
            work_sheet.cell(row=i + 2, column=3).value = ymin
            work_sheet.cell(row=i + 2, column=4).value = ymax
            work_sheet.cell(row=i + 2, column=5).value = xmin
            work_sheet.cell(row=i + 2, column=6).value = xmax
            work_sheet.cell(row=i + 2, column=7).value = head_body_length
            work_sheet.cell(row=i + 2, column=8).value = animal_weight

        work_sheet.cell(row=2, column=9).value = sheep_flag
        work_sheet.cell(row=2, column=10).value = cattle_flag

        work_book.save(json_path + "/" + filepath[:-5] + "_result_mask_label.xlsx")
        print("保存成功")
